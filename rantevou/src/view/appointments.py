from __future__ import annotations

import tkinter as tk
from tkinter import ttk, IntVar
from datetime import datetime, timedelta
from typing import Any
from itertools import chain
from collections import deque
import os
import sys
import webbrowser
import openpyxl
import subprocess

from .abstract_views import AppFrame

from ..controller.appointments_controller import AppointmentControl
from ..controller.logging import Logger
from ..controller.subscriber import SubscriberInterface
from ..controller.mailer import Mailer
from ..controller import get_config

from ..model.entities import Appointment, Customer

logger = Logger("AppointmentsTab")
cfg: dict[str, Any] = get_config()["view_settings"]
cfg["group_period"] = timedelta(hours=cfg["working_hours"] // cfg["rows"])
cfgb = get_config()["buttons"]


class AppointmentsTab(AppFrame, SubscriberInterface):

    appointments: list[Appointment]
    appointment_groups: dict[int, list[Appointment]]
    start_date: datetime
    group_period: timedelta
    end_date: datetime

    mail_panel: Grid

    # Shared data
    start_date = datetime.now().replace(
        hour=cfg["opening_hour"],
        minute=0,
        second=0,
        microsecond=0,
    )

    end_date = start_date + timedelta(
        days=cfg["columns"],
        hours=cfg["working_hours"],
    )

    group_period = cfg["group_period"]

    # appointment_groups = AppointmentControl().get_appointments_grouped_in_periods(
    #     start=start_date, period=timedelta(minutes=120)
    # )

    def __init__(self, root, *args, **kwargs):
        super().__init__(root, *args, **kwargs)
        SubscriberInterface.__init__(self)
        self.main_panel = Grid(self, AppointmentsTab.start_date)
        self.main_panel.pack(fill="both", expand=True)

    def subscriber_update(self):
        pass


class GridNavBar(ttk.Frame):
    move_left: ttk.Button
    move_right: ttk.Button

    def __init__(self, root: Grid, *args, **kwargs):
        super().__init__(root, *args, **kwargs)
        self.now = datetime.now()
        self.year = IntVar()
        self.month = IntVar()
        self.day = IntVar()
        self.set_date(timedelta(0))

        self.move_left = ttk.Button(
            self,
            text="Previous",
            command=lambda: self.set_date(timedelta(days=-1)) or root.move_left(1),
            style="low.TButton",
        )
        self.move_right = ttk.Button(
            self,
            text="Next",
            command=lambda: self.set_date(timedelta(days=1)) or root.move_right(1),
            style="low.TButton",
        )

        self.move_to_year = ttk.Entry(self, textvariable=self.year, width=4)
        self.move_to_month = ttk.Entry(self, textvariable=self.month, width=2)
        self.move_to_day = ttk.Entry(self, textvariable=self.day, width=2)

        for widget in [self.move_to_year, self.move_to_month, self.move_to_day]:
            widget.bind("<Key>", lambda x: x.keysym == "Return" and root.move_date(self.year, self.month, self.day))

        self.send_email_button = ttk.Button(self, text="@", command=self.send_email)
        self.print_button = ttk.Button(self, text="Print", command=self.print)
        self.export_excel = ttk.Button(self, text="Export", command=self.export_to_worksheet)

        self.move_right.pack(side=tk.RIGHT)
        self.move_left.pack(side=tk.RIGHT)
        self.move_to_day.pack(side=tk.LEFT)
        self.move_to_month.pack(side=tk.LEFT)
        self.move_to_year.pack(side=tk.LEFT)
        self.send_email_button.pack(side=tk.LEFT)
        self.print_button.pack(side=tk.LEFT)
        self.export_excel.pack(side=tk.LEFT)

    def set_date(self, delta: timedelta):
        self.now += delta
        self.year.set(self.now.year)
        self.month.set(self.now.month)
        self.day.set(self.now.day)

    def _get_appointments_from_entry(self) -> list[Appointment]:
        start_date = datetime(year=self.year.get(), month=self.month.get(), day=self.day.get())
        return AppointmentControl().get_appointments_from_to_date(start=start_date, end=start_date + timedelta(days=1))

    def send_email(self):
        appointments = self._get_appointments_from_entry()
        mailer = Mailer()
        mailer.send_email(appointments)

    def export_to_worksheet(self):
        appointments = self._get_appointments_from_entry()
        workbook = openpyxl.workbook.Workbook()
        workbook.create_sheet("Ραντεβού", 0)
        sheet = workbook["Ραντεβού"]

        for i, appointment in enumerate(appointments, 1):
            data = []
            data.extend(appointment.values)
            customer = appointment.customer
            if customer:
                data.extend(customer.values)

            for j, entry in enumerate(data, 1):
                sheet.cell(i, j, str(entry))

        workbook.save("tmp.xlsx")
        if sys.platform.lower().startswith("linux"):
            try:
                subprocess.run(["xdg-open", "tmp.xlsx"], check=True)
            except subprocess.CalledProcessError:
                try:
                    subprocess.run(["open", "tmp.xlsx"], check=True)
                except subprocess.CalledProcessError:
                    logger.log_error(f"Failed to open xlsx reader in platform {sys.platform}")

        elif sys.platform.lower().startswith("win"):
            try:
                subprocess.run(["start", "tmp.xlsx"], check=True)
            except subprocess.CalledProcessError:
                logger.log_error(f"Failed to open xlsx reader in platform {sys.platform}")

        elif sys.platform.lower().startswith("darwin"):
            try:
                subprocess.run(["open", "tmp.xlsx"], check=True)
            except subprocess.CalledProcessError:
                logger.log_error(f"Failed to open xlsx reader in platform {sys.platform}")

    def print(self):
        appointments = self._get_appointments_from_entry()
        buffer: list[str] = [f"<p>{self.day.get():02d}/{self.month.get():02d}/{self.year.get()}<br><table>"]
        for appointment in appointments:
            hour = appointment.date.hour
            minute = appointment.date.minute
            duration = int(appointment.duration.total_seconds() // 60)
            string = f"<tr><td>{hour:02d}:{minute:02d}</td><td>{duration:4}</td>"
            customer: Customer = appointment.customer
            name = "---"
            surname = "---"
            phone = "---"
            email = "---"
            if customer:
                _, name, surname, phone, email = (value or "---" for value in customer.values)
            string += f"<td>{name}</td><td>{surname}</td><td>{phone}</td><td>{email}</td><tr>"

            buffer.append(string)

        buffer.append("</p><script>print()</script>")
        filename = f"tmp{str(datetime.now())}.html"
        with open(filename, "w") as f:
            f.writelines(buffer)

        webbrowser.open(filename)
        self.after(10000, lambda: os.remove(filename))


class Grid(ttk.Frame):
    columns: list[GridColumn]
    navbar: GridNavBar
    start_date: datetime

    def __init__(self, root, start_date, *args, **kwargs):
        super().__init__(root, *args, **kwargs)
        self.start_date = start_date
        self.period_duration = cfg["group_period"]
        self.columns = []

        self.navbar = GridNavBar(self, *args, **kwargs)
        self.navbar.pack(fill="x")

        self.rowheader = GridVerticalHeader(self, start_date, self.period_duration, cfg["rows"])
        self.rowheader.pack(side=tk.LEFT, fill="both")

        for i in range(7):
            start_date = self.start_date + timedelta(days=i)
            end_date = start_date + timedelta(hours=cfg["working_hours"])
            start_index = AppointmentControl().get_index_from_date(start_date, self.start_date, self.period_duration)
            self.columns.append(
                GridColumn(
                    self,
                    start_date=start_date,
                    end_date=end_date,
                    start_index=start_index,
                    period_duration=self.period_duration,
                )
            )
            self.columns[i].pack(side=tk.LEFT, fill="both", expand=True, pady=10)

    def move_left(self, step=1):
        self.start_date -= timedelta(days=step)
        for column in self.columns:
            column.move_left(step)

    def move_right(self, step=1):
        self.start_date += timedelta(days=step)
        for column in self.columns:
            column.move_right(step)

    def move_date(self, year: IntVar, month: IntVar, day: IntVar):
        date = datetime(
            year=year.get(), month=month.get(), day=day.get(), hour=self.start_date.hour, minute=self.start_date.minute
        )
        diff = self.start_date - date
        if diff > timedelta(0):
            self.move_left(step=int(abs(diff).total_seconds() // (3600 * 24)))
        else:
            self.move_right(step=int(abs(diff).total_seconds() // (3600 * 24)))


class GridHeader(ttk.Label):
    date: datetime

    def __init__(self, root, date: datetime, *args, **kwargs):
        super().__init__(root, *args, **kwargs)
        self.date = date
        self.draw()

    def draw(self):
        self.config(text=self.date.strftime("%d/%m"))

    def move_left(self, step=1):
        self.date -= timedelta(days=step)
        self.draw()

    def move_right(self, step=1):
        self.date += timedelta(days=step)
        self.draw()


class GridVerticalHeader(ttk.Label):
    start: datetime
    interval: timedelta
    rows: int

    def __init__(self, master, start: datetime, interval: timedelta, rows: int, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.start = start
        self.interval = interval
        self.rows = rows
        for i in range(rows):
            from_time = (self.start + interval * i).strftime("%H:%M")
            to_time = (self.start + interval * (i + 1)).strftime("%H:%M")
            text = f"{from_time}-{to_time}"
            ttk.Label(self, text=text).pack(fill="both", expand=True)


class GridColumn(ttk.Frame):
    header: GridHeader
    rows: list[GridCell]
    start_date: datetime
    end_date: datetime
    period_duration: timedelta

    def __init__(
        self,
        root,
        start_date,
        end_date,
        start_index,
        period_duration,
        *args,
        **kwargs,
    ):
        super().__init__(root, *args, **kwargs)
        self.start_date = start_date
        self.end_date = end_date
        self.period_duration = period_duration
        self.start_index = start_index
        self.header = GridHeader(self, start_date)
        self.rows = []

        start = start_date
        i = 0
        while start < end_date:
            self.rows.append(GridCell(self, start_index + i, start, period_duration))
            i += 1
            start += period_duration

        self.header.pack(fill="x")
        for row in self.rows:
            row.pack(fill="both", expand=True, padx=3, pady=3)

    def move_left(self, step=1):
        self.header.move_left(step)
        for row in self.rows:
            row.move_left(step)

    def move_right(self, step=1):
        self.header.move_right(step)
        for row in self.rows:
            row.move_right(step)


class GridCell(ttk.Frame, SubscriberInterface):
    group_index: int
    period_start: datetime
    period_duration: timedelta
    current_focus = None

    def __init__(
        self,
        root,
        appointment_group_index: int,
        period_start: datetime,
        period_duration: timedelta,
        *args,
        **kwargs,
    ):
        super().__init__(root, *args, **kwargs)
        SubscriberInterface.__init__(self)

        self.group_index = appointment_group_index
        self.period_start = period_start
        self.period_duration = period_duration
        self.period_end = period_start + period_duration

        self.text = ttk.Label(self)
        self.text.pack(fill="both", expand=True, padx=3, pady=3)
        self.text.bind("<Button-1>", lambda x: self.show_in_sidepanel())
        self.config(borderwidth=1)

        self.cache: None | list = None
        self.draw()

    @property
    def appointments(self):
        result = AppointmentControl().get_appointments_from_to_date(
            self.period_start, self.period_start + timedelta(hours=2)
        )
        result.sort(key=lambda x: x.date)
        return result

    @property
    def times_between_appointments(self) -> list[tuple[datetime, timedelta]]:
        return AppointmentControl().get_time_between_appointments(
            start_date=self.period_start,
            end_date=self.period_end,
            minumum_free_period=timedelta(0),
        )

    @property
    def previous_period_appointments(self):
        appointments = AppointmentControl().get_appointments_from_to_date(
            self.period_start - timedelta(hours=2), self.period_start
        )
        appointments.sort(key=lambda x: x.date)
        return appointments

    @property
    def next_period_appointments(self):
        appointments = AppointmentControl().get_appointments_from_to_date(
            self.period_start + timedelta(hours=2), self.period_end + timedelta(hours=4)
        )
        appointments.sort(key=lambda x: x.date)
        return appointments

    def subscriber_update(self):
        self.cache = None
        self.draw()

    def draw(self):
        appointment_count = len(self.appointments)

        if appointment_count <= 1:
            self.style = "low.TLabel"
        elif appointment_count == 2:
            self.style = "medium.TLabel"
        elif appointment_count == 3:
            self.style = "high.TLabel"
        elif appointment_count > 3:
            self.style = "max.TLabel"
        self.text.config(style=self.style)

        text = f"Ραντεβού:{appointment_count}"
        self.text.config(text=text)

    def move_left(self, step):
        self.period_start -= timedelta(days=step)
        self.period_end -= timedelta(days=step)
        self.group_index = AppointmentControl().get_index_from_date(
            self.period_start, AppointmentsTab.start_date, self.period_duration
        )
        self.cache = None
        self.draw()

    def move_right(self, step):
        self.period_start += timedelta(days=step)
        self.period_end += timedelta(days=step)
        self.group_index = AppointmentControl().get_index_from_date(
            self.period_start, AppointmentsTab.start_date, self.period_duration
        )
        self.draw()
        self.cache = None

    def unfocus(self):
        """
        Αλλάζει το πλαίσιο γύρω από το κελί ώστε να δείχνει ανενεργό στον χρήστη
        """
        self.config(relief="flat")

    def show_focus(self):
        """
        Αλλάζει το πλαίσιο γύρω από το κελί ώστε να δείχνει ενεργό στον χρήστη
        """
        if GridCell.current_focus:
            GridCell.current_focus.unfocus()
        self.config(relief="raised")
        GridCell.current_focus = self

    def show_in_sidepanel(self):
        """
        Εντολή για την εμφανίση των ραντεβού στο sidepanel. Φτιάχνει μια λίστα
        από γεμάτες και άδειες περιόδους ώστε ο χρήστης με ένα κουμπί να προσθέτει
        ραντεβού στην ημερομηνία που θέλει. Ελέγχει τις προηγούμενες και επόμενες
        περιόδους ώστε να μην υπάρχει overlap.
        """

        self.show_focus()
        sidepanel = self.nametowidget(".!sidepanel")

        # if self.cache:
        #     logger.log_debug(f"{self.winfo_name()} showing cached version")
        #     return sidepanel.select_view("appointments", self, self.cache)

        min_duration = timedelta(minutes=cfg["minimum_appointment_duration"])

        # Εάν δεν υπάρχουν ραντεβού στην συγκεκριμένη περίοδο, κατασκευάζει μια
        # λίστα με άδεια ραντεβού
        if len(self.appointments) == 0:
            return sidepanel.select_view(
                "appointments",
                self,
                [
                    Appointment(
                        date=self.period_start + i * min_duration,
                        duration=min_duration,
                    )
                    for i in range(self.period_duration // min_duration)
                ],
            )

        # Κοιτάζει την προηγούμενη και επόμενη περίοδο. Εάν δεν υπάρχουν ραντεβού
        # τότε δημιουργεί πλασματικά ώστε να οριστεί η αρχή και το τέλος
        previous_appointments = self.previous_period_appointments
        next_appointments = self.next_period_appointments

        if len(previous_appointments) == 0:
            previous_appointments = [Appointment(date=self.period_start, duration=timedelta(0))]

        if len(next_appointments) == 0:
            next_appointments = [Appointment(date=self.period_end, duration=timedelta(0))]

        # Ενώνει όλα τα ραντεβού σε μια διπλή λίστα για πιο γρήγορα operations
        appointments: deque[Appointment] = deque(
            chain(
                previous_appointments,
                self.appointments,
                next_appointments,
            )
        )
        slots: list[Appointment] = []
        is_eligible = lambda a, b: a < self.period_end and b > self.period_start

        # Ανα ζεύγη, μέχρι να αδείασει η διπλή λίστα, τα βάζει στην τελική λίστα "slots",
        # με μορφή "Προηγούμενο ραντεβού | Άδειες περίοδοι 20 λεπτών ή λιγότερο | Επόμενο ραντεβού"
        while len(appointments) > 1:
            previous = appointments.popleft()
            next = appointments[0]

            # Η συνάρτηση is_eligible ελέγχει αν ο χρόνος του ραντεβού είναι εντός
            # της περιόδου που αφορά το συγκεκριμένο κελί
            if is_eligible(previous.date, previous.end_date):
                slots.append(previous)

            # diff -> ελεύθερος χρόνος μεταξύ 2 ραντεβού
            # full_appointments -> πόσα 20λεπτα ραντεβού χωράνε στο diff
            # remainer -> χρόνος λιγότερος των 20 λεπτών που απομένει
            diff = previous.time_between_appointments(next)
            full_appointments = diff // min_duration
            remainer = diff % min_duration

            for i in range(full_appointments):
                date = previous.end_date + min_duration * i
                duration = min_duration
                end_date = date + duration

                if is_eligible(date, end_date):
                    slots.append(
                        Appointment(
                            date=date,
                            duration=duration,
                        )
                    )

            if remainer > timedelta(0):
                date = next.date - remainer
                duration = remainer
                end_date = next.date

                if is_eligible(date, end_date):
                    slots.append(
                        Appointment(
                            date=date,
                            duration=duration,
                        )
                    )

        if appointments:
            last = appointments.pop()
            if is_eligible(last.end_date, last.end_date):
                slots.append(last)

        sidepanel.select_view("appointments", self, slots)
